from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, send_file
from flask_ngrok import run_with_ngrok
from datetime import datetime, timedelta, date
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import mysql.connector, io, json, logging
from mysql.connector import Error

# === Flask app setup ===
app = Flask(__name__)
app.config['SECRET_KEY'] = 'yallow'  # replace with a strong secret on production

run_with_ngrok(app)

# === DATABASE CONNECTION ===
def get_db_connection():
    """Return a new MySQL database connection."""
    return mysql.connector.connect(
        host='localhost',
        user='root',
        password='',
        database='db_ssc_database'
    )

# === CONTEXT PROCESSOR ===
@app.context_processor
def inject_user():
    # Provide username and user_role to templates
    return dict(username=session.get('username'), user_role=session.get('user_role'))

# === ROUTES ===
@app.route('/')
def redirect_home():
    return redirect(url_for('index'))

@app.route('/ssc_financial_management')
def index():
    return render_template('index.html')

@app.route('/administrator')
def administrator():
    return redirect(url_for('accounts'))

@app.context_processor
def inject_user():
    return dict(username=session.get('user_name'))

# === ACCOUNTS ===
@app.route('/accounts')
def accounts():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    # Include status column
    cursor.execute("SELECT id, username, email, role, password, status FROM tbl_ssc_accounts")
    accounts = cursor.fetchall()
    cursor.close()
    conn.close()
    return render_template('account_set_block.html', accounts=accounts)

# === ACCOUNT REGISTER ===
@app.route('/register', methods=['POST'])
def register():
    name = request.form.get('name')
    email = request.form.get('email')
    password = request.form.get('password')
    role = request.form.get('role')

    if not (name and email and password):
        flash('Please provide name, email and password.', 'register-error')
        return redirect(url_for('accounts'))

    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        # Check for duplicates
        cursor.execute("SELECT * FROM tbl_ssc_accounts WHERE username = %s OR email = %s", (name, email))
        existing_user = cursor.fetchone()

        if existing_user:
            flash('Username or Email already registered.', 'register-error')
        else:
            # Save account with force_change = 1
            cursor.execute("""
                INSERT INTO tbl_ssc_accounts (username, email, password, role, status, force_change)
                VALUES (%s, %s, %s, %s, %s, %s)
            """, (name, email, password, role, 'active', 1))
            conn.commit()
            flash('Registration successful! User must change password on first login.', 'success')
    except Exception as e:
        print("DB Error:", e)
        flash('Database error occurred.', 'register-error')
    finally:
        try:
            cursor.close()
            conn.close()
        except:
            pass

    return redirect(url_for('accounts'))

# ===== UPDATE PASSWORD ROUTE =====
@app.route('/update_password', methods=['POST'])
def update_password():
    if 'user_id' not in session:
        return jsonify({'status': 'error', 'message': 'Not logged in'})

    new_password = request.form.get('new_password')
    if not new_password:
        return jsonify({'status': 'error', 'message': 'Password required'})

    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute(
            "UPDATE tbl_ssc_accounts SET password=%s, force_change=0 WHERE id=%s",
            (new_password, session['user_id'])
        )
        conn.commit()
        return jsonify({'status': 'success'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})
    finally:
        cursor.close()
        conn.close()

# === Toggle Account Status (Enable / Disable)
@app.route('/toggle_account_status', methods=['POST'])
def toggle_account_status():
    account_id = request.form.get('account_id')
    new_status = request.form.get('status')

    if not account_id or new_status not in ['active', 'disabled']:
        return {"status": "error", "message": "Invalid request data!"}, 400

    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("UPDATE tbl_ssc_accounts SET status=%s WHERE id=%s", (new_status, account_id))
        conn.commit()
    except Exception as e:
        print("DB Error:", e)
        return {"status": "error", "message": "Database error occurred!"}, 500
    finally:
        try:
            cursor.close()
            conn.close()
        except:
            pass

    return {"status": "success", "message": f"Account {new_status} successfully!"}, 200

# === PROFILE PAGE ===
@app.route('/profile_page')
def profile_page():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    user_id = session['user_id']

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT id, username, email, role, status FROM tbl_ssc_accounts WHERE id=%s", (user_id,))
    user = cursor.fetchone()
    cursor.close()
    conn.close()

    return render_template('profile_block.html', user=user)

# === Verify Old Password ===
@app.route('/verify_old_password', methods=['POST'])
def verify_old_password():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    user_id = session['user_id']
    entered_pw = request.form['current_password']

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT password FROM tbl_ssc_accounts WHERE id=%s", (user_id,))
    stored_pw = cursor.fetchone()['password']
    cursor.close()
    conn.close()

    if entered_pw == stored_pw:
        session['old_password_verified'] = True
        return redirect(url_for('profile'))
    else:
        flash("Incorrect current password!", "danger")
        return redirect(url_for('profile_page'))


# === CHANGE PASSWORD ===
@app.route('/password_change', methods=['POST'])
def password_change():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    user_id = session['user_id']
    current_pw = request.form['current_password']
    new_pw = request.form['new_password']
    confirm_pw = request.form['confirm_password']

    if new_pw != confirm_pw:
        flash("New passwords do not match!", "danger")
        return redirect(url_for('profile_page'))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # Get existing password
    cursor.execute("SELECT password FROM tbl_ssc_accounts WHERE id=%s", (user_id,))
    stored_pw = cursor.fetchone()['password']

    # Check current password (plaintext version - adjust if hashed)
    if current_pw != stored_pw:
        flash("Current password is incorrect!", "danger")
        return redirect(url_for('profile_page'))

    # Update password
    cursor.execute("UPDATE tbl_ssc_accounts SET password=%s WHERE id=%s", (new_pw, user_id))
    conn.commit()
    cursor.close()
    conn.close()

    flash("Password successfully updated!", "success")
    return redirect(url_for('profile_page'))

# === TRANSACTION PAGE ADMIN ===
@app.route('/transaction_admin')
def transaction_admin():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM tbl_transaction_set_up ORDER BY id ASC")
    transactions = cursor.fetchall()
    cursor.close()
    conn.close()

    return render_template('transaction_admin_block.html', transactions=transactions)

# === TRANSACTION PAGE TREASURER ===
@app.route('/transaction')
def transaction():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM tbl_transaction_set_up ORDER BY id ASC")
    transactions = cursor.fetchall()
    cursor.close()
    conn.close()

    username = session.get('username', 'Guest')
    return render_template('transaction_block.html', transactions=transactions, username=username)

# === ADD TRANSACTION FEE ===
@app.route('/transaction/add', methods=['POST'])
def add_transaction():
    description = request.form.get('description')
    total_fee = request.form.get('total-fee')
    school_year = request.form.get('school_year')
    semester = request.form.get('semester')

    if not (description and total_fee and school_year and semester):
        flash('Missing fields for transaction.', 'error')
        return redirect(url_for('transaction'))

    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("""
            INSERT INTO tbl_transaction_set_up 
            (description, total_fee, school_year, semester)
            VALUES (%s, %s, %s, %s)
        """, (description, total_fee, school_year, semester))

        conn.commit()
        flash('Transaction added successfully.', 'success')
    except Exception as e:
        print("Error adding transaction:", e)
        flash('Failed to add transaction.', 'error')
    finally:
        cursor.close()
        conn.close()

    return redirect(url_for('transaction'))

# === UPDATE TRANSACTION FEE ===
@app.route('/transaction/update_fee', methods=['POST'])
def update_fee():
    data = request.get_json()
    tx_id = data.get('id')
    total_fee = data.get('total_fee')

    if not tx_id or not total_fee:
        return jsonify({'status': 'error', 'message': 'Missing data'})

    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "UPDATE tbl_transaction_set_up SET total_fee = %s WHERE id = %s",
            (total_fee, tx_id)
        )
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({'status': 'success'})
    except Exception as e:
        print("Error updating fee:", e)
        return jsonify({'status': 'error', 'message': str(e)})


# === REMOVE TRANSACTION FEE ===
@app.route('/transaction/remove_fee', methods=['POST'])
def remove_fee():
    data = request.get_json()
    tx_id = data.get('id')

    if not tx_id:
        return jsonify({'status': 'error', 'message': 'Missing transaction ID'})

    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM tbl_transaction_set_up WHERE id = %s", (tx_id,))
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({'status': 'success'})
    except Exception as e:
        print("Error removing transaction:", e)
        return jsonify({'status': 'error', 'message': str(e)})


# === View Locker Set Up ===
@app.route('/locker_set_up')
def locker_set_up():
    username = session.get('username', 'Guest')
    lockers = []

    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        # Fetch lockers exactly as stored in the database
        cursor.execute("""
            SELECT id, locker_letter, locker_number, status
            FROM tbl_lockers
            ORDER BY locker_letter ASC,
                    CAST(locker_number AS UNSIGNED) ASC
        """)
        lockers = cursor.fetchall()

    except Exception as e:
        print("Error fetching lockers:", e)
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

    return render_template('locker_set_block.html', username=username, lockers=lockers)

# === Add Locker ===
@app.route('/locker/add', methods=['POST'])
def add_locker():
    locker_letter = request.form.get('locker-letter', '').strip().upper()
    locker_number = request.form.get('locker-number', '').strip()
    # Default status is 'Available'
    status = request.form.get('locker-status', 'Available')
    # Compose default locker name
    locker_name = f"LOCKER {locker_letter}-{locker_number}"

    if not locker_letter or not locker_number:
        return jsonify(status='error', message='Letter and number are required')

    try:
        locker_number = str(int(locker_number))  # Ensure numeric
    except ValueError:
        return jsonify(status='error', message='Invalid locker number')

    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO tbl_lockers (locker_letter, locker_number, status, locker_name) VALUES (%s, %s, %s, %s)",
            (locker_letter, locker_number, status, locker_name)
        )
        conn.commit()
        return jsonify(status='success', id=cursor.lastrowid, name=locker_name)
    except Exception as e:
        return jsonify(status='error', message=str(e))
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

# === Update Locker ===
@app.route('/locker/update_name', methods=['POST'])
def update_locker():
    data = request.get_json()
    locker_id = data.get('id')
    locker_letter = data.get('letter', '').strip().upper()
    locker_number = str(data.get('number')).strip()
    locker_name = data.get('locker_name', f"LOCKER {locker_letter}-{locker_number}")
    status = data.get('status', 'Available')  # Optional, can be ignored

    if not locker_id or not locker_letter or not locker_number:
        return jsonify(status='error', message='Invalid locker data')

    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "UPDATE tbl_lockers SET locker_letter=%s, locker_number=%s, locker_name=%s, status=%s WHERE id=%s",
            (locker_letter, locker_number, locker_name, status, locker_id)
        )
        conn.commit()
        return jsonify(status='success')
    except Exception as e:
        return jsonify(status='error', message=str(e))
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

# === Remove Locker ===
@app.route('/locker/remove', methods=['POST'])
def remove_locker():
    data = request.get_json()
    locker_id = data.get('id')

    if not locker_id:
        return jsonify(status='error', message='Invalid locker ID')

    conn = None
    cursor = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM tbl_lockers WHERE id = %s", (locker_id,))
        conn.commit()
        return jsonify(status='success')
    except Exception as e:
        print("Error removing locker:", e)
        return jsonify(status='error', message=str(e))
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

# === LOGIN / PANELS ===
@app.route('/enter_officer_panel', methods=['GET', 'POST'])
def ssc_login_panel():
    if request.method == 'POST':
        return redirect(url_for('ssc_login_page'))
    return render_template('ssc_login_panel.html')

@app.route('/ssc_login_page')
def ssc_login_page():
    return render_template('ssc_officers_block.html')

@app.route('/enter_viewing_panel', methods=['GET', 'POST'])
def ssc_viewing_panel():
    if request.method == 'POST':
        return redirect(url_for('ssc_viewing_page'))
    return render_template('ssc_viewing_panel.html')

@app.route('/ssc_viewing_page')
def ssc_viewing_page():
    return render_template('ssc_viewing_panel.html')

@app.route('/enter_department_panel', methods=['GET', 'POST'])
def ssc_department_panel():
    if request.method == 'POST':
        return redirect(url_for('ssc_department_page'))
    return render_template('ssc_department_panel.html')

@app.route('/ssc_department_page')
def ssc_department_page():
    return render_template('ssc_department_panel.html')

# === UNIVERSAL LOGIN & ACTIVITY LOGGING (Single Row, Update Only Activity) ===
@app.before_request
def log_user_activity():
    user_id = session.get('user_id')
    if not user_id:
        return  # Skip if not logged in

    # Skip static files and certain routes
    excluded_routes = ['static', 'favicon', 'ssc_login_page', 'logout']
    if any(request.path.startswith(f'/{r}') for r in excluded_routes):
        return

    # Get page/action name
    page_name = (request.endpoint or 'unknown').split('.')[-1]
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    new_activity = f"{timestamp} - {page_name}"
    ip_address = request.remote_addr

    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        # Find the latest normal log for this user
        cursor.execute("""
            SELECT id, activity
            FROM tbl_users_log
            WHERE user_id = %s AND logout_type = 'normal'
            ORDER BY login_time DESC
            LIMIT 1
        """, (user_id,))
        last_log = cursor.fetchone()

        if last_log:
            # Update the existing row: only activity (and optionally IP)
            updated_activity = new_activity  # replace or append if you want
            cursor.execute("""
                UPDATE tbl_users_log
                SET activity = %s, ip_address = %s
                WHERE id = %s
            """, (updated_activity, ip_address, last_log['id']))
        else:
            # Insert a new row at first login
            cursor.execute("""
                INSERT INTO tbl_users_log (user_id, ip_address, activity, logout_type)
                VALUES (%s, %s, %s)
            """, (user_id, ip_address, new_activity))

        conn.commit()

    except Exception as e:
        print("Activity logging error:", e)
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

# ==============================
# ===== LOGIN ROUTE =====
@app.route('/login', methods=['POST'])
def officers_login():
    login_input = request.form.get('email')  # username or email
    password = request.form.get('password')

    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        # Determine if input is email or username
        if '@' in login_input and '.' in login_input:
            query = "SELECT * FROM tbl_ssc_accounts WHERE email = %s"
        else:
            query = "SELECT * FROM tbl_ssc_accounts WHERE username = %s"

        cursor.execute(query, (login_input,))
        user = cursor.fetchone()

        if not user:
            flash("Username or email not found.", "danger")
            return redirect(url_for('ssc_login_page'))

        if user.get('status', 'active') == 'disabled':
            flash("Your account has been disabled.", "danger")
            return redirect(url_for('ssc_login_page'))

        if user['password'] != password:
            flash("Incorrect password.", "danger")
            return redirect(url_for('ssc_login_page'))

        # Store session info
        session['user_id'] = user['id']
        session['username'] = user['username']
        session['email'] = user['email']
        session['user_role'] = user.get('role', '').lower()

        # === Log user login ===
        ip_address = request.remote_addr
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        cursor.execute("""
            INSERT INTO tbl_users_log (user_id, ip_address, activity, logout_type, login_time)
            VALUES (%s, %s, %s, 'normal', %s)
        """, (user['id'], ip_address, f"{timestamp} - officers_login", timestamp))
        conn.commit()
        session['current_log_id'] = cursor.lastrowid

        # ===== Role-based handling =====
        role = session['user_role']

        # If user must change password
        if user.get('force_change', 0):
            return redirect(url_for('change_password'))

        # Redirect based on role
        if role == 'operation_activities':
            return redirect(url_for('main_menu'))
        elif role == 'financial_activities':
            return redirect(url_for('auditor_notes'))
        elif role == 'admin':
            return redirect(url_for('administrator'))
        else:
            return redirect(url_for('ssc_login_page'))

    except Exception as e:
        print("Login Error:", e)
        flash("An error occurred during login.", "danger")
        return redirect(url_for('ssc_login_page'))
    finally:
        if cursor: cursor.close()
        if conn: conn.close()

# ===== CHANGE PASSWORD PAGE =====
@app.route('/change_password', methods=['GET', 'POST'])
def change_password():
    if 'user_id' not in session:
        flash("Please log in first.", "danger")
        return redirect(url_for('ssc_login_page'))

    if request.method == 'POST':
        new_password = request.form.get('new_password')
        confirm_password = request.form.get('confirm_password')

        if not new_password or not confirm_password:
            flash("Please fill out both fields.", "danger")
            return redirect(url_for('change_password'))

        if new_password != confirm_password:
            flash("Passwords do not match.", "danger")
            return redirect(url_for('change_password'))

        if len(new_password) < 8:
            flash("Password must be at least 8 characters.", "danger")
            return redirect(url_for('change_password'))

        # Update DB
        conn = get_db_connection()
        cursor = conn.cursor()
        try:
            cursor.execute(
                "UPDATE tbl_ssc_accounts SET password=%s, force_change=0 WHERE id=%s",
                (new_password, session['user_id'])
            )
            conn.commit()
            flash("Password updated successfully!", "success")

            # Redirect based on role after password change
            role = session.get('user_role', '')
            if role == 'operation_activities':
                return redirect(url_for('main_menu'))
            elif role == 'financial_activities':
                return redirect(url_for('auditor_notes'))
            elif role == 'admin':
                return redirect(url_for('administrator'))
            else:
                return redirect(url_for('ssc_login_page'))

        finally:
            cursor.close()
            conn.close()

    return render_template('change_password.html', username=session['username'])

# ==============================
# === VIEWING LOGIN ROUTE ===
# ==============================
@app.route('/viewing_login', methods=['POST'])
def viewing_login():
    login_input = request.form.get('email')  # username or email
    password = request.form.get('password')

    conn = None
    cursor = None

    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        # Determine if input is email or username
        if '@' in login_input and '.' in login_input:
            query = "SELECT * FROM tbl_ssc_accounts WHERE email = %s"
        else:
            query = "SELECT * FROM tbl_ssc_accounts WHERE username = %s"

        cursor.execute(query, (login_input,))
        user = cursor.fetchone()

        # Account not found
        if not user:
            flash("Username or email not found.", "login-error")
            return redirect(url_for('ssc_viewing_page'))

        # Disabled account
        if user.get('status', 'active') == 'disabled':
            flash("Your account has been disabled.", "login-error")
            return redirect(url_for('ssc_viewing_page'))

        # Wrong password
        if user['password'] != password:
            flash("Incorrect password.", "login-error")
            return redirect(url_for('ssc_viewing_page'))

        # Store session
        session['user_id'] = user['id']
        session['username'] = user['username']
        session['email'] = user['email']
        session['user_role'] = user.get('role', '').lower()

        # Force password change
        if user.get('force_change', 0):
            return redirect(url_for('change_password_viewing'))

        # Successful login redirection rules
        role = session['user_role']

        if role == 'admin':
            return redirect(url_for('transaction'))

        elif role == 'vice_admin':
            return redirect(url_for('vice_admin'))

        # Default viewer role
        return redirect(url_for('ssc_viewing_page'))

    except Exception as e:
        print("Viewing Login Error:", e)
        flash("An error occurred during login.", "login-error")
        return redirect(url_for('ssc_viewing_page'))

    finally:
        if cursor: cursor.close()
        if conn: conn.close()

# ===== CHANGE PASSWORD VIEWING PAGE =====
@app.route('/change_password_viewing', methods=['GET', 'POST'])
def change_password_viewing():
    if 'user_id' not in session:
        flash("Please log in first.", "danger")
        return redirect(url_for('ssc_login_page'))

    if request.method == 'POST':
        new_password = request.form.get('new_password')
        confirm_password = request.form.get('confirm_password')

        if not new_password or not confirm_password:
            flash("Please fill out both fields.", "danger")
            return redirect(url_for('change_password'))

        if new_password != confirm_password:
            flash("Passwords do not match.", "danger")
            return redirect(url_for('change_password'))

        if len(new_password) < 8:
            flash("Password must be at least 8 characters.", "danger")
            return redirect(url_for('change_password'))

        # Update DB
        conn = get_db_connection()
        cursor = conn.cursor()
        try:
            cursor.execute(
                "UPDATE tbl_ssc_accounts SET password=%s, force_change=0 WHERE id=%s",
                (new_password, session['user_id'])
            )
            conn.commit()
            flash("Password updated successfully!", "success")
            return redirect(url_for('vice_admin'))
        finally:
            cursor.close()
            conn.close()

    return render_template('change_password.html', username=session['username'])

# === MAIN PAGE ===
@app.route('/main_menu')
def main_menu():
    if 'username' not in session:
        return redirect(url_for('ssc_login_page'))
    return render_template('menu_block.html', username=session['username'])

# === VICE ADMIN ===
@app.route('/vice_admin')
def vice_admin():
    if 'username' not in session:
        return redirect(url_for('ssc_login_page'))
    return render_template('vice_admin_block.html', username=session['username'])

# === MEMBERSHIPS/PAYMENT ===
@app.route('/membership', methods=['GET', 'POST'])
def membership_page():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # Fetch transactions
    cursor.execute("SELECT * FROM tbl_transaction_set_up ORDER BY id ASC")
    transactions = cursor.fetchall()

    # Next serial number
    cursor.execute("SELECT MAX(serial_number) AS last_serial FROM tbl_scc_memberships")
    last_serial = cursor.fetchone()['last_serial'] or 0
    next_serial_formatted = f"{last_serial + 1:04d}"

    if request.method == 'POST':
        data = request.get_json()
        students = data.get('students', [])
        if not students:
            return jsonify({"success": False, "message": "No students provided."}), 400

        try:
            for s in students:
                # Fetch transaction details
                cursor.execute("SELECT * FROM tbl_transaction_set_up WHERE id=%s", (s['transaction_id'],))
                transaction = cursor.fetchone()

                # Force status = 'Approved'
                status = 'Approved'

                logging.info(f"Inserting membership for {s['student_number']} with status: {status}")

                # Insert membership
                cursor.execute("""
                    INSERT INTO tbl_scc_memberships
                    (serial_number, student_number, member_type, surname, first_name, middle_name,
                     transferee, department, course, level, date_joined, school_year, semester,
                     officer, organization, position, status)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, (
                    None if s.get('serial_number') in ['N/A', None] else s['serial_number'],
                    s['student_number'],
                    s['member_type'],
                    s['surname'],
                    s['firstname'],
                    s.get('middlename', ''),
                    s.get('transferee', ''),
                    s['department'],
                    s['course'],
                    s['level'],
                    datetime.now(),
                    s.get('school_year', 'N/A'),
                    s.get('semester', 'N/A'),
                    s.get('officer', 'No'),
                    s.get('organization') or 'N/A',
                    s.get('position') or 'N/A',
                    status
                ))
                membership_id = cursor.lastrowid

                # Insert transaction record
                if transaction:
                    cursor.execute("""
                        INSERT INTO tbl_transactions (membership_id, date, description)
                        VALUES (%s, %s, %s)
                    """, (membership_id, datetime.now().date(), transaction['description']))

                # Insert cash inflow if action_type is membership
                if s.get('action_type') == 'membership':
                    cursor.execute("""
                        INSERT INTO tbl_operating_activities_recieved
                        (membership_id, student_number, surname, first_name, middle_name, department, course,
                         description_fee, transaction_fee, date_created, school_year, semester,
                         officer, organization, position, status)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    """, (
                        membership_id,
                        s['student_number'],
                        s['surname'],
                        s['firstname'],
                        s.get('middlename', ''),
                        s['department'],
                        s['course'],
                        s.get('description_fee', 'N/A'),
                        transaction['total_fee'] if transaction else 0,
                        datetime.now(),
                        s.get('school_year', 'N/A'),
                        s.get('semester', 'N/A'),
                        s.get('officer', 'No'),
                        s.get('organization') or 'N/A',
                        s.get('position') or 'N/A',
                        status
                    ))

            conn.commit()
            return jsonify({"success": True})

        except Exception as e:
            conn.rollback()
            logging.error(f"Error saving membership: {str(e)}")
            return jsonify({"success": False, "message": str(e)}), 500

        finally:
            cursor.close()
            conn.close()

    cursor.close()
    conn.close()
    return render_template(
        'memberships.html',
        transactions=transactions,
        next_serial_formatted=next_serial_formatted,
        username=session.get('username', 'Guest')
    )


# === PAYMENT ONLY ROUTE ===
@app.route('/payment', methods=['POST'])
def payment_page_post():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        data = request.get_json()
        payments = data.get('students', [])
        if not payments:
            return jsonify({"success": False, "message": "No payments provided."}), 400

        messages = []

        for p in payments:
            if p.get('action_type') != 'payment':
                continue

            # Check duplicate payment
            cursor.execute("""
                SELECT COUNT(*) AS count
                FROM tbl_operating_activities_recieved
                WHERE student_number = %s AND description_fee = %s
            """, (p['student_number'], p.get('description_fee', 'N/A')))
            if cursor.fetchone()['count'] > 0:
                messages.append(f"{p['surname']}, {p['firstname']} has already paid {p.get('description_fee','N/A')}.")
                continue

            # Fetch transaction details
            cursor.execute("SELECT * FROM tbl_transaction_set_up WHERE id = %s", (p['transaction_id'],))
            transaction = cursor.fetchone()

            # Get membership_id
            cursor.execute("SELECT id FROM tbl_scc_memberships WHERE student_number = %s", (p['student_number'],))
            membership_id = cursor.fetchone()
            membership_id = membership_id['id'] if membership_id else None

            status = 'Approved'
            logging.info(f"Inserting payment for {p['student_number']} with status: {status}")

            # Insert transaction record
            if transaction:
                cursor.execute("""
                    INSERT INTO tbl_transactions (membership_id, date, description)
                    VALUES (%s, %s, %s)
                """, (membership_id, datetime.now().date(), transaction['description']))

            # Insert payment record
            cursor.execute("""
                INSERT INTO tbl_operating_activities_recieved
                (membership_id, student_number, surname, first_name, middle_name, department, course,
                 description_fee, transaction_fee, date_created, school_year, semester,
                 officer, organization, position, status)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (
                membership_id,
                p['student_number'],
                p['surname'],
                p['firstname'],
                p.get('middlename', ''),
                p['department'],
                p['course'],
                p.get('description_fee', 'N/A'),
                transaction['total_fee'] if transaction else 0,
                datetime.now(),
                p.get('school_year', 'N/A'),
                p.get('semester', 'N/A'),
                p.get('officer', 'No'),
                p.get('organization') or 'N/A',
                p.get('position') or 'N/A',
                status
            ))

        conn.commit()
        return jsonify({"success": True, "messages": messages})

    except Exception as e:
        conn.rollback()
        logging.error(f"Error saving payment: {str(e)}")
        return jsonify({"success": False, "message": str(e)}), 500

    finally:
        cursor.close()
        conn.close()
        
# === GET STUDENT API (Auto-Fetch, Membership + Students) ===
@app.route('/api/get_student/<student_number>', methods=['GET'])
def api_get_student(student_number):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # Try finding in memberships table first
    cursor.execute("""
        SELECT 
            student_number, 
            surname, 
            first_name, 
            middle_name, 
            department, 
            course, 
            level, 
            member_type, 
            transferee, 
            serial_number
        FROM tbl_scc_memberships
        WHERE student_number = %s
    """, (student_number,))
    student = cursor.fetchone()

    # If not found in memberships, check students table
    if not student:
        cursor.execute("""
            SELECT 
                student_number, 
                surname, 
                first_name, 
                middle_name, 
                department, 
                course, 
                level,
                'N/A' AS member_type, 
                'No' AS transferee, 
                'N/A' AS serial_number
            FROM tbl_students
            WHERE student_number = %s
        """, (student_number,))
        student = cursor.fetchone()

    cursor.close()
    conn.close()

    if student:
        # Clean nulls to safe defaults
        for key in ['surname', 'first_name', 'middle_name', 'department', 'course', 'level', 'member_type', 'transferee', 'serial_number']:
            if student.get(key) is None:
                student[key] = 'N/A'

        return jsonify({"success": True, "student": student})
    else:
        return jsonify({"success": False, "message": "Student not found."}), 404

# === UPDATE STUDENT INFO FOR BOTH TABLES ===
@app.route('/update_student_payment', methods=['POST'])
def update_student_payment():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        data = request.get_json()
        student_number = data.get('student_number')

        # --- Update tbl_operating_activities_recieved ---
        cursor.execute("""
            UPDATE tbl_operating_activities_recieved
            SET surname=%s,
                first_name=%s,
                middle_name=%s,
                department=%s,
                course=%s,
                level=%s,
                school_year=%s,
                semester=%s,
                officer=%s,
                organization=%s,
                position=%s
            WHERE student_number=%s
        """, (
            data.get('surname'),
            data.get('firstname'),
            data.get('middle_name'),
            data.get('department'),
            data.get('course'),
            data.get('level'),
            data.get('school_year'),
            data.get('semester'),
            data.get('officer'),
            data.get('organization'),
            data.get('position'),
            student_number
        ))

        # --- Update tbl_scc_memberships ---
        cursor.execute("""
            UPDATE tbl_scc_memberships
            SET surname=%s,
                first_name=%s,
                middle_name=%s,
                department=%s,
                course=%s,
                level=%s,
                school_year=%s,
                semester=%s,
                officer=%s,
                organization=%s,
                position=%s
            WHERE student_number=%s
        """, (
            data.get('surname'),
            data.get('firstname'),
            data.get('middle_name'),
            data.get('department'),
            data.get('course'),
            data.get('level'),
            data.get('school_year'),
            data.get('semester'),
            data.get('officer'),
            data.get('organization'),
            data.get('position'),
            student_number
        ))

        conn.commit()
        return jsonify({"success": True})

    except Exception as e:
        conn.rollback()
        return jsonify({"success": False, "message": str(e)}), 500

    finally:
        cursor.close()
        conn.close()

# Show pending registration page
@app.route('/member_registration')
def member_registration():
    if 'user_id' not in session:
        flash("Please log in to access settings.", "login-error")
        return redirect(url_for('index'))
    
    username = session.get('username', 'Guest')
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # Fetch transactions
    cursor.execute("SELECT * FROM tbl_transaction_set_up ORDER BY id ASC")
    transactions = cursor.fetchall()

    # Fetch pending registrations
    cursor.execute("SELECT * FROM tbl_scc_memberships WHERE status='Pending' ORDER BY id ASC")
    pending_registrations = cursor.fetchall()

    cursor.close()
    conn.close()

    # Format serial numbers as 4-digit strings
    for p in pending_registrations:
        if p.get('serial_number') is not None:
            p['serial_number'] = str(p['serial_number']).zfill(4)
        else:
            p['serial_number'] = "0000"

    return render_template(
        'pending_registration_block.html',
        transactions=transactions,
        pending_registrations=pending_registrations,
        username=username
    )


# API: fetch pending registrations (for dynamic refresh)
@app.route('/api/pending_registrations')
def api_pending_registrations():
    if 'user_id' not in session:
        return jsonify({"success": False, "pending": []})

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM tbl_scc_memberships WHERE status='Pending' ORDER BY id ASC")
    pending = cursor.fetchall()
    cursor.close()
    conn.close()

    for p in pending:
        if p.get('serial_number') is not None:
            p['serial_number'] = str(p['serial_number']).zfill(4)
        else:
            p['serial_number'] = "0000"

    return jsonify({"success": True, "pending": pending})


# API: confirm student membership
@app.route('/api/confirm_membership', methods=['POST'])
def api_confirm_membership():
    if 'user_id' not in session:
        return jsonify({"success": False, "message": "Not logged in"}), 403

    data = request.get_json()
    student = data.get('student')
    if not student:
        return jsonify({"success": False, "message": "No student data"}), 400

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        # Generate next serial number
        cursor.execute("SELECT MAX(serial_number) AS last_serial FROM tbl_scc_memberships")
        last_serial = cursor.fetchone()['last_serial'] or 0
        next_serial = str(int(last_serial) + 1).zfill(4)

        # Fetch transaction details
        cursor.execute("SELECT * FROM tbl_transaction_set_up WHERE id=%s", (student['transaction_id'],))
        transaction = cursor.fetchone()

        # Update existing membership
        cursor.execute("""
            UPDATE tbl_scc_memberships
            SET 
                member_type=%s,
                surname=%s,
                first_name=%s,
                middle_name=%s,
                department=%s,
                course=%s,
                date_joined=%s,
                school_year=%s,
                semester=%s,
                status='Active',
                officer=%s,
                organization=%s,
                position=%s
            WHERE student_number=%s
        """, (
            student['member_type'],
            student['surname'],
            student['firstname'],
            student.get('middle_name',''),
            student['department'],
            student['course'],
            datetime.now(),
            student.get('school_year','N/A'),
            student.get('semester','N/A'),
            student.get('officer','No'),
            student.get('organization','N/A'),
            student.get('position','N/A'),
            student['student_number']
        ))

        # If you need the membership_id for inserting into operating activities:
        cursor.execute("SELECT id FROM tbl_scc_memberships WHERE student_number=%s", (student['student_number'],))
        membership_id = cursor.fetchone()['id']

        # Insert into operating activities
        if transaction:
            cursor.execute("""
                INSERT INTO tbl_operating_activities_recieved
                (membership_id, student_number, surname, first_name, middle_name, department, course,
                 description_fee, transaction_fee, date_created, school_year, semester,
                 officer, organization, position)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (
                membership_id,
                student['student_number'],
                student['surname'],
                student['firstname'],
                student.get('middle_name',''),
                student['department'],
                student['course'],
                transaction['description'],
                transaction['total_fee'],
                datetime.now(),
                student.get('school_year','N/A'),
                student.get('semester','N/A'),
                student.get('officer','No'),
                student.get('organization','N/A'),
                student.get('position','N/A')
            ))

        # Remove from pending by updating status
        cursor.execute("UPDATE tbl_scc_memberships SET status='Active' WHERE student_number=%s", (student['student_number'],))

        conn.commit()
        return jsonify({"success": True, "serial_number": next_serial})

    except Exception as e:
        conn.rollback()
        return jsonify({"success": False, "message": str(e)}), 500
    finally:
        cursor.close()
        conn.close()

# === REGISTRATION / MEMBERSHIPS ===
@app.route('/registration', methods=['GET', 'POST'])
def registration():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # Determine next serial number
    cursor.execute("SELECT MAX(serial_number) AS last_serial FROM tbl_scc_memberships")
    last_serial = cursor.fetchone()['last_serial'] or 0
    next_serial_formatted = f"{last_serial + 1:04d}"

    if request.method == 'POST':
        data = request.get_json()
        student_list = data.get('students', [])
        if not student_list:
            return jsonify({"success": False, "message": "No students provided."}), 400

        try:
            for s in student_list:
                cursor.execute("""
                    INSERT INTO tbl_scc_memberships
                    (serial_number, student_number, member_type, surname, first_name, middle_name,
                     transferee, officer, organization, position, department, course, level,
                     status, school_year, semester, date_joined)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, (
                    None if s['serial_number'] == 'N/A' else s['serial_number'],
                    s['student_number'],
                    s['member_type'],
                    s['surname'],
                    s['firstname'],
                    s.get('middlename', ''),
                    s.get('transferee', 'no'),
                    s.get('officer', 'no'),
                    s.get('organization') if s.get('organization') not in [None, '', 'null'] else 'N/A',
                    s.get('position') if s.get('position') not in [None, '', 'null'] else 'N/A',
                    s['department'],
                    s['course'],
                    s['level'],
                    'Pending',  # pending status
                    s.get('school_year', 'N/A'),
                    s.get('semester', 'N/A'),
                    datetime.now()
                ))

            conn.commit()
            return jsonify({"success": True, "message": "Student(s) added with Pending status."})

        except Exception as e:
            conn.rollback()
            return jsonify({"success": False, "message": str(e)}), 500

        finally:
            cursor.close()
            conn.close()

    # GET request → render registration form
    cursor.close()
    conn.close()
    username = session.get('username', 'Guest')
    return render_template(
        'membership_registration_block.html',
        next_serial_formatted=next_serial_formatted,
        username=username
    )

# === EXPENSES: Add multiple expenses ===
@app.route('/expenses', methods=['GET', 'POST'])
def expenses():
    username = session.get('username', 'Guest')
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    if request.method == 'POST':
        data = request.get_json()  # Expecting a list of expense objects

        if not isinstance(data, list) or len(data) == 0:
            conn.close()
            return jsonify({"status": "error", "message": "No data provided"}), 400

        saved_count = 0
        last_inserted_id = None

        for item in data:
            category = item.get('category')
            expense_date = item.get('expense_date')
            description = item.get('description') or ""   # ← FIX: avoid NULL breaking duplicates
            si_or_no = item.get('si_or_no')
            amount = item.get('amount')

            # Skip invalid entries
            if not all([category, expense_date, si_or_no, amount]):
                continue

            # Validate amount
            try:
                amount = float(amount)
            except (ValueError, TypeError):
                continue

            # Duplicate check
            cursor.execute("""
                SELECT COUNT(*) AS count FROM tbl_operating_activities_payments
                WHERE category=%s AND expense_date=%s AND description=%s 
                AND si_or_no=%s AND amount=%s
            """, (category, expense_date, description, si_or_no, amount))

            if cursor.fetchone()['count'] > 0:
                continue  # Skip duplicate

            # Insert
            cursor.execute("""
                INSERT INTO tbl_operating_activities_payments
                (category, expense_date, description, si_or_no, amount, created_date)
                VALUES (%s, %s, %s, %s, %s, CURDATE())
            """, (category, expense_date, description, si_or_no, amount))

            last_inserted_id = cursor.lastrowid
            saved_count += 1

        conn.commit()
        conn.close()
        return jsonify({
            "status": "success",
            "message": f"{saved_count} expense(s) saved successfully.",
            "payment_id": last_inserted_id
        })

    # GET request but **without querying the DB** (as you requested)
    conn.close()
    return render_template("procurement_block.html", username=username)
    
# === EDIT HISTORY ===
@app.route('/edit_expense', methods=['POST'])
def edit_expense():
    data = request.get_json()

    category = data.get('category')
    old_data = data.get('old_data')
    new_data = data.get('new_data')
    edited_by = data.get('edited_by') or session.get('username', 'Unknown')

    if not category or not old_data or not new_data:
        return jsonify({'status':'error', 'message':'Missing required data'}), 400

    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        # --- 1. Update main expenses table ---
        update_query = """
            UPDATE tbl_expenses
            SET expense_date=%s, description=%s, si_or_no=%s, amount=%s
            WHERE category=%s AND payment_id=%s
        """
        cursor.execute(update_query, (
            new_data['date'],
            new_data['description'],
            new_data['si_or_no'],
            new_data['amount'],
            category,
            old_data.get('payment_id')  # Ensure your storedData has payment_id
        ))

        # --- 2. Insert into edit history table ---
        history_query = """
            INSERT INTO tbl_expense_history
            (payment_id, old_date, old_description, old_si_or_no, old_amount,
             new_date, new_description, new_si_or_no, new_amount,
             edited_by)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """
        cursor.execute(history_query, (
            old_data.get('payment_id'),
            old_data['date'],
            old_data['description'],
            old_data['si_or_no'],
            old_data['amount'],
            new_data['date'],
            new_data['description'],
            new_data['si_or_no'],
            new_data['amount'],
            edited_by
        ))

        conn.commit()
        cursor.close()
        conn.close()

        return jsonify({'status':'success', 'message':'Expense updated and logged successfully!'})

    except Exception as e:
        return jsonify({'status':'error', 'message': str(e)}), 500

# === UPDATE EXPENSES WITH HISTORY ===
@app.route('/update_expense', methods=['POST'])
def update_expense():
    data = request.get_json()  # Expecting a list of expense updates

    if not isinstance(data, list) or len(data) == 0:
        return jsonify({"status": "error", "message": "Invalid data"}), 400

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    updated_count = 0
    updated_rows = []

    try:
        for item in data:
            expense_id = item.get('id')  # <-- Updated to use 'id'
            category = item.get('category')
            new_date = item.get('expense_date')
            new_desc = item.get('description')
            new_si = item.get('si_or_no')
            new_amount = float(item.get('amount'))
            edited_by = "current_user"  # Replace with session username if available

            if not all([expense_id, category, new_date, new_desc, new_si, new_amount is not None]):
                continue

            # --- Fetch old values from tbl_operating_activities_payments ---
            cursor.execute("SELECT expense_date, description, si_or_no, amount FROM tbl_operating_activities_payments WHERE id=%s", (expense_id,))
            old = cursor.fetchone()
            if not old:
                continue

            old_date = old['expense_date']
            old_desc = old['description']
            old_si = old['si_or_no']
            old_amount = float(old['amount'])

            # --- Insert old + new values into history ---
            cursor.execute("""
                INSERT INTO tbl_expense_history
                (payment_id, old_date, old_description, old_si_or_no, old_amount, new_date, new_description, new_si_or_no, new_amount, edited_by)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (expense_id, old_date, old_desc, old_si, old_amount, new_date, new_desc, new_si, new_amount, edited_by))

            # --- Update main expense ---
            cursor.execute("""
                UPDATE tbl_operating_activities_payments
                SET category=%s, expense_date=%s, description=%s, si_or_no=%s, amount=%s
                WHERE id=%s
            """, (category, new_date, new_desc, new_si, new_amount, expense_id))

            updated_count += 1
            updated_rows.append({
                "id": expense_id,
                "category": category,
                "expense_date": new_date,
                "description": new_desc,
                "si_or_no": new_si,
                "amount": new_amount
            })

        conn.commit()
        if updated_count > 0:
            return jsonify({
                "status": "success",
                "message": f"{updated_count} expense(s) updated successfully.",
                "updated_rows": updated_rows
            })
        else:
            return jsonify({"status": "error", "message": "No valid expenses were updated."})

    except Exception as e:
        conn.rollback()
        return jsonify({"status": "error", "message": str(e)})
    finally:
        cursor.close()
        conn.close()

# === EDIT HISTORY ===
@app.route('/expense_history/<int:payment_id>')
def expense_history(payment_id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    username = session.get('username', 'Guest')

    # Fetch all history for this payment
    cursor.execute("""
        SELECT id,
               old_date,
               old_description,
               old_si_or_no,
               old_amount,
               new_date,
               new_description,
               new_si_or_no,
               new_amount,
               edited_by,
               edited_at
        FROM tbl_expense_history
        WHERE payment_id = %s
        ORDER BY edited_at DESC
    """, (payment_id,))
    
    history = cursor.fetchall()
    conn.close()

    return render_template("expense_history.html", history=history, username=username)

# === EDIT HISTORY PROCUREMENT ===
@app.route('/expense_history_procurement/<int:payment_id>')
def expense_history_procurement(payment_id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    username = session.get('username', 'Guest')

    # Fetch all history for this payment
    cursor.execute("""
        SELECT id,
               old_date,
               old_description,
               old_si_or_no,
               old_amount,
               new_date,
               new_description,
               new_si_or_no,
               new_amount,
               edited_by,
               edited_at
        FROM tbl_expense_history
        WHERE payment_id = %s
        ORDER BY edited_at DESC
    """, (payment_id,))
    
    history = cursor.fetchall()
    conn.close()

    return render_template("expense_history_procurement.html", history=history, username=username)

# === RENTAL LOCKER PAGE ===
@app.route('/rental', methods=['GET', 'POST'])
def rental_page():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    username = session.get('username', 'Guest')

    # Fetch lockers
    cursor.execute("SELECT locker_name, locker_letter, locker_number, status FROM tbl_lockers ORDER BY locker_letter, locker_number")
    lockers = cursor.fetchall()

    # Group lockers by letter
    lockers_by_letter = {}
    for locker in lockers:
        lockers_by_letter.setdefault(locker['locker_letter'], []).append(locker)

    # Fetch active rentals
    cursor.execute("""
        SELECT r.id, r.locker_letter, r.locker_number,
               CONCAT(r.first_name,' ',r.middle_name,' ',r.surname) AS name,
               r.student_number, r.department, r.course, r.due_date, r.status
        FROM tbl_locker_rentals r
        WHERE r.status='Active'
        ORDER BY r.due_date ASC
    """)
    rentals = cursor.fetchall()

    if request.method == 'POST':
        student_number = request.form.get('student_number')
        surname = request.form.get('surname')
        firstname = request.form.get('first_name')
        middlename = request.form.get('middle_name', '')
        department = request.form.get('department')
        course = request.form.get('course')
        contact_number = request.form.get('contactNumber')
        locker_select = request.form.get('lockerSelect')

        # Debug
        print("Form submission:", request.form)

        if not all([student_number, surname, firstname, department, course, contact_number, locker_select]):
            flash("Please fill in all required fields.", "error")
            return redirect(url_for('rental_page'))

        # Parse locker letter + number
        try:
            locker_letter, locker_number = locker_select.split('-')
            locker_number = int(locker_number)
        except:
            flash("Invalid locker selection.", "error")
            return redirect(url_for('rental_page'))

        # Check duplicate rental
        cursor.execute("SELECT * FROM tbl_locker_rentals WHERE student_number=%s AND status='Active'", (student_number,))
        if cursor.fetchone():
            flash(f"Student {firstname} {surname} already has an active locker.", "error")
            return redirect(url_for('rental_page'))

        # Check locker availability
        cursor.execute("SELECT status FROM tbl_lockers WHERE locker_letter=%s AND locker_number=%s", (locker_letter, locker_number))
        locker = cursor.fetchone()
        if not locker or locker['status'] == 'Occupied':
            flash(f"Locker {locker_letter}-{locker_number} is occupied.", "error")
            return redirect(url_for('rental_page'))

        # Save rental
        now = datetime.now()
        date_rented = now.strftime("%Y-%m-%d %H:%M:%S")
        due_date = (now + timedelta(days=30)).strftime("%Y-%m-%d %H:%M:%S")

        try:
            cursor.execute("""
                INSERT INTO tbl_locker_rentals
                (locker_letter, locker_number, student_number, surname, first_name, middle_name,
                 department, course, contact_number, date_rented, due_date, status)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'Active')
            """, (locker_letter, locker_number, student_number, surname, firstname, middlename,
                  department, course, contact_number, date_rented, due_date))
            
            cursor.execute("UPDATE tbl_lockers SET status='Occupied' WHERE locker_letter=%s AND locker_number=%s",
                           (locker_letter, locker_number))
            conn.commit()
            flash(f"Locker {locker_letter}-{locker_number} successfully rented to {firstname} {surname}!", "success")
        except Exception as e:
            conn.rollback()
            print("Error saving rental:", e)
            flash(f"An error occurred: {e}", "error")

        return redirect(url_for('rental_page'))

    cursor.close()
    conn.close()

    return render_template('locker_rent.html', lockers_by_letter=lockers_by_letter, lockers=lockers, rentals=rentals, username=username)

# === API: Fetch Student & Check Rental Payment / Active Locker ===
@app.route('/api/check_rental_payment/<student_number>', methods=['GET'])
def check_rental_payment(student_number):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        # 1. Get student info
        cursor.execute("SELECT * FROM tbl_scc_memberships WHERE student_number=%s", (student_number,))
        student = cursor.fetchone()
        if not student:
            return jsonify({"success": False, "message": "Student not found."})

        # 2. Check rental locker payment
        cursor.execute("""
            SELECT * FROM tbl_operating_activities_recieved
            WHERE student_number=%s AND LOWER(description_fee) LIKE %s
            ORDER BY date_created DESC LIMIT 1
        """, (student_number, '%rental locker%'))
        payment = cursor.fetchone()

        # 3. Check if student already has an active locker
        cursor.execute("""
            SELECT locker_letter, locker_number 
            FROM tbl_locker_rentals
            WHERE student_number=%s AND status='Active'
        """, (student_number,))
        active_locker = cursor.fetchone()

        # Combine letter + number for exact locker identifier
        if active_locker:
            student['active_locker'] = f"{active_locker['locker_letter']}{active_locker['locker_number']}"
        else:
            student['active_locker'] = None

        # 4. Return responses based on status
        if student['active_locker']:
            return jsonify({
                "success": False,
                "message": f"Student already has an active locker (# {student['active_locker']}).",
                "student": student
            })

        if not payment:
            return jsonify({
                "success": False,
                "message": "Student has not paid the rental locker fee.",
                "student": student
            })

        # 5. If everything is okay
        return jsonify({
            "success": True,
            "student": student,
            "payment": payment
        })

    finally:
        cursor.close()
        conn.close()

# === RENEW ===
@app.route('/renew/<int:rental_id>', methods=['GET', 'POST'])
def renew_rental(rental_id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    # Get rental info
    cursor.execute("SELECT * FROM tbl_locker_rentals WHERE id=%s", (rental_id,))
    rental = cursor.fetchone()

    if request.method == 'POST':
        fee = request.form.get('transaction_fee')
        description = request.form.get('description_fee', 'Locker Fee')  # default description

        # Validate fee
        if not fee or fee.strip() == '':
            flash("Transaction fee is required!")
            cursor.close()
            conn.close()
            return redirect(url_for('renew_rental', rental_id=rental_id))

        # Save fee into tbl_operating_activities_recieved
        cursor.execute("""
            INSERT INTO tbl_operating_activities_recieved 
            (membership_id, student_number, surname, first_name, middle_name, department, course, description_fee, transaction_fee) 
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (
            rental['id'],                     # membership_id
            rental['student_number'],
            rental['surname'],
            rental['first_name'],
            rental['middle_name'],
            rental['department'],
            rental['course'],
            description,
            fee
        ))

        # Renew locker due date AND update status to 'Renewed'
        cursor.execute("""
            UPDATE tbl_locker_rentals 
            SET due_date = DATE_ADD(due_date, INTERVAL 30 DAY), 
                status = 'Renewed'
            WHERE id=%s
        """, (rental_id,))

        # --- ADD TRANSACTION RECORD ---
        cursor.execute("""
            INSERT INTO tbl_transactions (membership_id, date, description)
            VALUES (%s, %s, %s)
        """, (rental['id'], datetime.now().date(), description))
        # -----------------------------

        conn.commit()
        cursor.close()
        conn.close()
        flash("Transaction fee saved, locker renewed, and transaction recorded successfully!")
        return redirect(url_for('student_locker', rental_id=rental_id))

    cursor.close()
    conn.close()
    return render_template('renew_released_block.html', rental=rental)

# === RELEASE LOCKER ===
@app.route('/release/<int:rental_id>', methods=['POST'])
def release_rental(rental_id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # Get locker number for this rental
    cursor.execute("SELECT locker_number FROM tbl_locker_rentals WHERE id=%s", (rental_id,))
    rental = cursor.fetchone()

    if rental:
        locker_number = rental['locker_number']
        # Update rental status to Released
        cursor.execute("UPDATE tbl_locker_rentals SET status='Released' WHERE id=%s", (rental_id,))
        # Set locker status to Available
        cursor.execute("UPDATE tbl_lockers SET status='Available' WHERE locker_number=%s", (locker_number,))
        conn.commit()

    cursor.close()
    conn.close()

    flash("Locker successfully released!", "warning")
    return redirect(url_for('student_locker'))  # Go back to rentals page

# === STUDENT LOCKER ===
@app.route('/student_locker')
def student_locker():
    username = session.get('username', 'Guest')
    
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # Fetch ALL locker rentals
    cursor.execute("""
        SELECT 
            id,
            locker_letter,
            locker_number,
            student_number,
            surname,
            first_name,
            middle_name,
            department,
            course,
            contact_number,
            date_rented,
            due_date,
            status
        FROM tbl_locker_rentals
        ORDER BY date_rented DESC
    """)
    rentals = cursor.fetchall()

    # Get unique departments
    departments = sorted(list({r['department'] for r in rentals}))

    # Get courses by department
    courses_by_dept = {}
    for r in rentals:
        dept = r['department']
        course = r['course']
        if dept not in courses_by_dept:
            courses_by_dept[dept] = set()
        courses_by_dept[dept].add(course)
    # Convert sets to sorted lists
    courses_by_dept = {k: sorted(list(v)) for k, v in courses_by_dept.items()}

    cursor.close()
    conn.close()

    return render_template(
        "student_locker_block.html",
        username=username,
        rentals=rentals,
        departments=departments,
        courses_by_dept=courses_by_dept
    )

# === STUDENT PAYMENT RECORD ===
@app.route('/student_payment_record')
def student_payment_record():
    username = session.get('username', 'Guest')
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # --- Fetch all students including serial_number, ordered numerically ---
    cursor.execute("""
        SELECT 
            serial_number,
            student_number,
            surname,
            first_name,
            middle_name,
            department,
            course,
            level,
            school_year,
            semester,
            officer,
            organization,
            position
        FROM tbl_scc_memberships
        ORDER BY CAST(serial_number AS UNSIGNED) ASC
    """)
    students = cursor.fetchall()

    # --- Add formatted serial number ---
    for s in students:
        s["formatted_serial"] = str(s["serial_number"]).zfill(4) if s.get("serial_number") else "N/A"

        # Normalize fields
        s["officer"] = str(s.get("officer") or "no").lower()
        s["organization"] = str(s.get("organization") or "")
        s["position"] = str(s.get("position") or "")

    # --- Fetch unique departments, courses, levels, SY, semester ---
    cursor.execute("SELECT DISTINCT department FROM tbl_scc_memberships ORDER BY department ASC")
    departments = [row['department'] for row in cursor.fetchall()]

    cursor.execute("SELECT DISTINCT course FROM tbl_scc_memberships ORDER BY course ASC")
    courses = [row['course'] for row in cursor.fetchall()]

    cursor.execute("SELECT DISTINCT level FROM tbl_scc_memberships ORDER BY level ASC")
    year_levels = [row['level'] for row in cursor.fetchall()]

    cursor.execute("SELECT DISTINCT school_year FROM tbl_scc_memberships ORDER BY school_year DESC")
    school_year = [row['school_year'] for row in cursor.fetchall()]

    cursor.execute("SELECT DISTINCT semester FROM tbl_scc_memberships ORDER BY semester DESC")
    semester = [row['semester'] for row in cursor.fetchall()]

    cursor.execute("SELECT DISTINCT officer FROM tbl_scc_memberships ORDER BY officer ASC")
    officers = [str(row['officer']).lower() for row in cursor.fetchall()]

    # --- NEW: Fetch unique organizations ---
    cursor.execute("""
        SELECT DISTINCT organization 
        FROM tbl_scc_memberships
        WHERE organization IS NOT NULL AND organization <> ''
        ORDER BY organization ASC
    """)
    organizations = [row['organization'] for row in cursor.fetchall()]

    # --- Set current SY and Semester ---
    current_sy = school_year[0] if school_year else None
    current_sem = semester[0] if semester else None

    cursor.close()
    conn.close()

    return render_template(
        "student_payment_block.html",
        students=students,
        departments=departments,
        courses=courses,
        year_levels=year_levels,
        school_year=school_year,
        semester=semester,
        officers=officers,
        organizations=organizations,   # ← SEND TO TEMPLATE
        username=username,
        current_sy=current_sy,
        current_sem=current_sem
    )

# === VIEW INDIVIDUAL STUDENT INFO + PAYMENTS (Paid & Unpaid) ===
@app.route('/student_info/<student_number>')
def student_info(student_number):
    username = session.get('username', 'Guest')
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # --- Fetch student details ---
    cursor.execute("""
        SELECT *
        FROM tbl_scc_memberships
        WHERE student_number = %s
    """, (student_number,))
    student = cursor.fetchone()

    # --- Fetch distinct school years and semesters first ---
    cursor.execute("""
        SELECT DISTINCT school_year
        FROM tbl_operating_activities_recieved
        WHERE student_number = %s
        ORDER BY school_year DESC
    """, (student_number,))
    school_years = [row['school_year'] for row in cursor.fetchall()]
    latest_school_year = school_years[0] if school_years else ''

    cursor.execute("""
        SELECT DISTINCT semester
        FROM tbl_operating_activities_recieved
        WHERE student_number = %s
        ORDER BY semester ASC
    """, (student_number,))
    semesters = [row['semester'] for row in cursor.fetchall()]
    latest_semester = semesters[-1] if semesters else ''

    # --- Fetch all transactions with stored status and assign default SY/Semester if missing ---
    cursor.execute("""
        SELECT t.id, t.description, t.total_fee,
               IFNULL(p.total_paid, 0) AS paid_fee,
               (t.total_fee - IFNULL(p.total_paid, 0)) AS balance,
               COALESCE(p.status, 'Unpaid') AS status,
               p.date_paid,
               COALESCE(p.school_year, %s) AS school_year,
               COALESCE(p.semester, %s) AS semester
        FROM tbl_transaction_set_up t
        LEFT JOIN (
            SELECT description_fee, 
                   SUM(transaction_fee) AS total_paid,
                   MAX(date_created) AS date_paid,
                   MAX(status) AS status,
                   student_number,
                   school_year,
                   semester
            FROM tbl_operating_activities_recieved
            WHERE student_number = %s
            GROUP BY description_fee, student_number, school_year, semester
        ) p ON t.description = p.description_fee
        ORDER BY t.id ASC
    """, (latest_school_year, latest_semester, student_number))
    transactions = cursor.fetchall()

    # --- Compute totals ---
    total_paid = sum(float(t['paid_fee']) for t in transactions)
    total_balance = sum(float(t['balance']) for t in transactions)

    cursor.close()
    conn.close()

    return render_template(
        'student_info.html',
        student=student,
        transactions=transactions,
        total_paid=total_paid,
        total_balance=total_balance,
        username=username,
        school_years=school_years,
        semesters=semesters,
        latest_school_year=latest_school_year,
        latest_semester=latest_semester
    )

# === UPDATE INDIVIDUAL TRANSACTION (AJAX) ===
@app.route('/update_transaction', methods=['POST'])
def update_transaction():
    data = request.get_json()
    transaction_id = data.get('id')
    description = data.get('description')
    total_fee = data.get('total_fee')
    paid_fee = data.get('paid_fee')
    status = data.get('status')
    date_paid = data.get('date_paid')

    if not transaction_id:
        return jsonify({"success": False, "message": "Transaction ID missing"}), 400

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        # Update transaction setup
        cursor.execute("""
            UPDATE tbl_transaction_set_up
            SET description = %s, total_fee = %s
            WHERE id = %s
        """, (description, total_fee, transaction_id))

        # Check if a payment record exists
        cursor.execute("""
            SELECT COUNT(*) FROM tbl_operating_activities_recieved
            WHERE description_fee = %s AND student_number = (
                SELECT student_number FROM tbl_transaction_set_up WHERE id=%s
            )
        """, (description, transaction_id))
        exists = cursor.fetchone()[0]

        if exists > 0:
            # Update existing payment record
            cursor.execute("""
                UPDATE tbl_operating_activities_recieved
                SET transaction_fee = %s, status = %s, date_created = %s
                WHERE description_fee = %s AND student_number = (
                    SELECT student_number FROM tbl_transaction_set_up WHERE id=%s
                )
            """, (paid_fee, status, date_paid or None, description, transaction_id))
        else:
            # Insert new payment record if paid_fee > 0
            cursor.execute("SELECT student_number FROM tbl_transaction_set_up WHERE id=%s", (transaction_id,))
            student_number = cursor.fetchone()[0]
            if paid_fee > 0:
                cursor.execute("""
                    INSERT INTO tbl_operating_activities_recieved
                    (student_number, description_fee, transaction_fee, status, date_created)
                    VALUES (%s, %s, %s, %s, %s)
                """, (student_number, description, paid_fee, status, date_paid or None))

        conn.commit()
        return jsonify({"success": True, "message": "Transaction updated successfully"})
    except Exception as e:
        conn.rollback()
        return jsonify({"success": False, "message": str(e)}), 500
    finally:
        cursor.close()
        conn.close()

# === STUDENT PAID RECORDS PAGE ===
@app.route('/payment_records')
def payment_records():
    username = session.get('username', 'Guest')

    # --- Check if user is logged in ---
    if 'user_id' not in session:
        flash("Please log in to access payment records.", "login-error")
        return redirect(url_for('index'))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # --- Fetch all students ---
    cursor.execute("""
        SELECT student_number,
               CONCAT(first_name, ' ', COALESCE(middle_name,''), ' ', surname) AS name,
               department,
               course,
               school_year,
               semester
        FROM tbl_scc_memberships
        ORDER BY surname ASC
    """)
    students = cursor.fetchall()

    # --- Fetch all transactions ---
    cursor.execute("SELECT id, description, total_fee FROM tbl_transaction_set_up ORDER BY id ASC")
    transactions = cursor.fetchall()

    # Convert Decimal to float
    for t in transactions:
        t['total_fee'] = float(t['total_fee'])

    # --- Fetch all payments ---
    cursor.execute("""
        SELECT student_number,
               description_fee,
               SUM(transaction_fee) AS total_paid
        FROM tbl_operating_activities_recieved
        GROUP BY student_number, description_fee
    """)
    payments = cursor.fetchall()

    # --- Build lookup dictionary: student_number -> transaction_id -> paid_amount ---
    payment_lookup = {}
    for p in payments:
        student_no = p['student_number']
        trans_id = next((t['id'] for t in transactions if t['description'] == p['description_fee']), None)
        if trans_id:
            payment_lookup.setdefault(student_no, {})[trans_id] = float(p['total_paid'])

    # --- Extract unique filters ---
    departments = sorted({s['department'] for s in students})
    courses = sorted({s['course'] for s in students})
    school_years = sorted({s['school_year'] for s in students}, reverse=True)  # latest first
    semesters = sorted({s['semester'] for s in students})

    # --- Determine latest/current school year & semester ---
    current_school_year = school_years[0] if school_years else ''
    current_semester = semesters[-1] if semesters else ''

    cursor.close()
    conn.close()

    return render_template(
        'student_paid_block.html',
        students=students,
        transactions=transactions,
        payment_lookup=payment_lookup,
        departments=departments,
        courses=courses,
        school_years=school_years,
        semesters=semesters,
        username=username,
        current_school_year=current_school_year,  # for auto-select in dropdown
        current_semester=current_semester         # for auto-select in dropdown
    )

# === CASH INFLOWS
@app.route('/cash_flow')
def cash_flow_page():
    if 'user_id' not in session:
        flash("Please log in to access the finance reports.", "login-error")
        return redirect(url_for('index'))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        # ✅ Fetch all real student cash inflow data
        cursor.execute("""
            SELECT 
                id,
                DATE_FORMAT(date_created, '%Y-%m-%d') AS date,
                student_number,
                surname,
                first_name,
                middle_name,
                department,
                course,
                school_year,
                semester,
                description_fee AS description,
                transaction_fee AS payment
            FROM tbl_operating_activities_recieved
            ORDER BY date_created DESC
        """)
        transactions = cursor.fetchall()

        # ✅ Summary Totals
        cursor.execute("""
            SELECT IFNULL(SUM(transaction_fee),0) AS today_total 
            FROM tbl_operating_activities_recieved 
            WHERE DATE(date_created) = CURDATE()
        """)
        today_total = cursor.fetchone()['today_total']

        cursor.execute("""
            SELECT IFNULL(SUM(transaction_fee),0) AS week_total 
            FROM tbl_operating_activities_recieved 
            WHERE YEARWEEK(date_created,1) = YEARWEEK(CURDATE(),1)
        """)
        week_total = cursor.fetchone()['week_total']

        cursor.execute("""
            SELECT IFNULL(SUM(transaction_fee),0) AS month_total 
            FROM tbl_operating_activities_recieved 
            WHERE MONTH(date_created) = MONTH(CURDATE()) 
            AND YEAR(date_created) = YEAR(CURDATE())
        """)
        month_total = cursor.fetchone()['month_total']

        cursor.execute("""
            SELECT IFNULL(SUM(transaction_fee),0) AS total_all 
            FROM tbl_operating_activities_recieved
        """)
        total_all = cursor.fetchone()['total_all']

        # ✅ Fetch fee descriptions from fee setup table
        cursor.execute("SELECT description FROM tbl_transaction_set_up ORDER BY description ASC")
        fee_descriptions = [row['description'] for row in cursor.fetchall()]

    except Exception as e:
        flash(f"Database error: {str(e)}", "error")
        transactions = []
        today_total = week_total = month_total = total_all = 0
        fee_descriptions = []

    finally:
        cursor.close()
        conn.close()

    return render_template(
        'cash_report.html',
        transactions=transactions,
        today_total=today_total,
        week_total=week_total,
        month_total=month_total,
        total_all=total_all,
        fee_descriptions=fee_descriptions  # ✅ added
    )

# === EXPORT EXCEL VIA DESCRIPTION FEE ===
@app.route("/export_cash_inflows")
def export_cash_inflows():
    selected_fee = request.args.get("fee", "")

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    sql = """
        SELECT 
            DATE_FORMAT(date_created, '%Y-%m-%d') AS date,
            student_number,
            surname,
            first_name,
            middle_name,
            department,
            course,
            description_fee AS description,
            transaction_fee AS payment
        FROM tbl_operating_activities_recieved
    """

    params = []

    if selected_fee:
        sql += " WHERE description_fee = %s"
        params.append(selected_fee)

    sql += " ORDER BY date_created DESC"

    cursor.execute(sql, params)
    rows = cursor.fetchall()

    cursor.close()
    conn.close()

    # ✅ Create Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Cash Inflows"

    # ✅ Column Headers (MATCH PAGE)
    headers = [
        "#", "Date", "Student Number", "Student Name",
        "Department", "Course", "Description", "Amount (₱)"
    ]
    ws.append(headers)

    # ✅ Style definitions
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    # ✅ Borders
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # ✅ Header background
    header_fill = PatternFill(start_color="e2e8f0", end_color="e2e8f0", fill_type="solid")

    # ✅ Apply header styles
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = bold
        cell.alignment = center
        cell.border = border
        cell.fill = header_fill

    # ✅ Insert data rows
    for index, r in enumerate(rows, start=1):
        full_name = f"{r['surname']}, {r['first_name']} {r['middle_name']}".strip()

        row_data = [
            index,
            r["date"],
            r["student_number"],
            full_name,
            r["department"],
            r["course"],
            r["description"],
            float(r["payment"])
        ]

        ws.append(row_data)

    # ✅ Style data rows
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=8):
        for cell in row:
            cell.border = border
            cell.alignment = left

        # ✅ Center index column (#)
        row[0].alignment = center

        # ✅ Amount column formatting (₱ 0.00)
        row[7].number_format = '₱#,##0.00'

    # ✅ Auto column width
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter

        for cell in col:
            if cell.value:
                length = len(str(cell.value))
                if length > max_length:
                    max_length = length

        ws.column_dimensions[column].width = max_length + 3

    # ✅ Final Output
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Cash_Inflow_{selected_fee if selected_fee else 'All_Fees'}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# === AUDITOR ===
@app.route('/auditor_page')
def auditor_notes():
    username = session.get('username', 'Guest')
    if 'user_id' not in session:
        flash("Please log in to access settings.", "login-error")
        return redirect(url_for('index'))
    return render_template('auditor_menu_block.html', username=username)

# === BREAKDOWN NOTES ===
@app.route('/breakdown_notes')
def breakdown_notes():
    if 'user_id' not in session:
        flash("Please log in to access financial summary.", "login-error")
        return redirect(url_for('index'))

    username = session.get('username', 'Guest')
    filter_date = request.args.get('date')  # format: YYYY-MM-DD

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # --- Inflows ---
    inflow_query = """
        SELECT description_fee AS description,
               COUNT(*) AS quantity,
               transaction_fee AS fee,
               SUM(transaction_fee) AS total_amount
        FROM tbl_operating_activities_recieved
    """
    params = []
    if filter_date:
        inflow_query += " WHERE DATE(transaction_date) = %s"
        params.append(filter_date)
    inflow_query += " GROUP BY description_fee, transaction_fee"

    cursor.execute(inflow_query, params)
    received_inflows = cursor.fetchall() or []

    # --- Outflows ---
    outflow_query = """
        SELECT p.id,
               p.category,
               p.description,
               p.si_or_no,
               p.amount,
               p.expense_date,
               EXISTS(
                   SELECT 1
                   FROM tbl_expense_history h
                   WHERE h.payment_id = p.id
               ) AS has_history
        FROM tbl_operating_activities_payments p
    """
    if filter_date:
        outflow_query += " WHERE DATE(p.expense_date) = %s"
    outflow_query += " ORDER BY p.expense_date DESC, p.category"

    cursor.execute(outflow_query, params)
    payment_details = cursor.fetchall() or []

    conn.close()

    display_label = f"Transactions on {filter_date}" if filter_date else "All Transactions"

    # Calculate totals
    total_inflows = sum([row['total_amount'] or 0 for row in received_inflows])
    total_outflows = sum([row['amount'] or 0 for row in payment_details])
    net_balance = total_inflows - total_outflows

    return render_template(
        'transaction_summary.html',
        username=username,
        display_label=display_label,
        received_inflows=received_inflows,
        payment_details=payment_details,
        total_inflows=total_inflows,
        total_outflows=total_outflows,
        net_balance=net_balance,
        filter_date=filter_date
    )

# === FINANCIAL DASHBOARD ===
@app.route('/financial_report')
def financial_report():
    if 'user_id' not in session:
        flash("Please log in to access reports.", "login-error")
        return redirect(url_for('index'))

    username = session.get('username', 'Guest')
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        # Fetch unique transaction dates from inflows table
        cursor.execute("""
            SELECT DISTINCT DATE(date_created) AS date
            FROM tbl_operating_activities_recieved
            ORDER BY date_created DESC
        """)
        transaction_dates = cursor.fetchall()

    except Exception as e:
        flash(f"Database error: {str(e)}", "error")
        transaction_dates = []

    finally:
        cursor.close()
        conn.close()

    return render_template(
        'finance_report_block.html',
        username=username,
        transaction_dates=transaction_dates
    )

# === TRANSACTION DETAILS ===
@app.route('/transaction_details')
def transaction_details():
    if 'user_id' not in session:
        flash("Please log in to view transactions.", "login-error")
        return redirect(url_for('index'))

    username = session.get('username', 'Guest')
    trans_date = request.args.get('trans_date') or date.today().strftime('%Y-%m-%d')

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        print("Fetching transactions for:", trans_date)

        # --- Cash Inflows (tbl_operating_activities_recieved) ---
        cursor.execute("""
            SELECT description_fee AS description,
                   transaction_fee AS amount
            FROM tbl_operating_activities_recieved
            WHERE DATE(date_created) = %s
            ORDER BY date_created DESC
        """, (trans_date,))
        inflows = cursor.fetchall()
        print("Inflows found:", len(inflows))

        # --- Cash Outflows (tbl_operating_activities_payments) ---
        cursor.execute("""
            SELECT category AS description,
                   amount
            FROM tbl_operating_activities_payments
            WHERE DATE(created_date) = %s
            ORDER BY created_date DESC
        """, (trans_date,))
        outflows = cursor.fetchall()
        print("Outflows found:", len(outflows))

        display_label = f"Transactions for {trans_date}"

    except Exception as e:
        print("Database error:", str(e))
        flash(f"Database error: {str(e)}", "error")
        inflows = []
        outflows = []
        display_label = "Error"

    finally:
        cursor.close()
        conn.close()

    return render_template(
        'transaction_details.html',
        inflows=inflows,
        outflows=outflows,
        display_label=display_label,
        username=username
    )

# === ACCOUNT TRACKING ===
@app.route('/account_tracking')
def account_tracking():
    if 'user_id' not in session:
        flash("Please log in to access settings.", "login-error")
        return redirect(url_for('index'))

    username = session.get('username', 'Guest')

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # Fetch users to monitor
    cursor.execute("""
        SELECT id, username, email, role
        FROM tbl_ssc_accounts
        WHERE role IN ('operation_activities', 'financial_activities')
    """)
    users = cursor.fetchall()

    # Fetch logs
    user_ids = [str(u['id']) for u in users] if users else []
    logs = []

    if user_ids:
        format_ids = ",".join(user_ids)
        filter_user = request.args.get('filter_user')
        filter_date = request.args.get('filter_date')

        sql = f"""
            SELECT l.*, a.username, a.email, a.role
            FROM tbl_users_log l
            JOIN tbl_ssc_accounts a ON l.user_id = a.id
            WHERE l.user_id IN ({format_ids})
        """

        params = []
        if filter_user:
            sql += " AND l.user_id = %s"
            params.append(filter_user)
        if filter_date:
            sql += " AND DATE(l.login_time) = %s"
            params.append(filter_date)

        sql += " ORDER BY l.login_time DESC"
        cursor.execute(sql, tuple(params))
        logs = cursor.fetchall()

    cursor.close()
    conn.close()

    return render_template(
        'account_tracking_block.html',
        username=username,
        users=users,
        logs=logs
    )

# === ABOUT PAGE ===
@app.route('/about')
def about_us_page():
    username = session.get('username', 'Guest')
    if 'user_id' not in session:
        flash("Please log in to access settings.", "login-error")
        return redirect(url_for('index'))
    return render_template('about_blocks.html', username=username)

# === SETTINGS PAGE ===
@app.route('/settings')
def settings():
    username = session.get('username', 'Guest')
    if 'user_id' not in session:
        flash("Please log in to access settings.", "login-error")
        return redirect(url_for('index'))
    return render_template('setting_blocks.html', username=username)

# === LOGOUT ===
@app.route('/logout')
def logout():
    log_id = session.get('current_log_id')
    if log_id:
        try:
            conn = get_db_connection()
            cursor = conn.cursor()
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            cursor.execute("""
                UPDATE tbl_users_log
                SET logout_time = %s, logout_type = 'manual'
                WHERE id = %s
            """, (timestamp, log_id))
            conn.commit()
        except Exception as e:
            print("Logout logging error:", e)
        finally:
            if cursor: cursor.close()
            if conn: conn.close()

    session.clear()
    return redirect(url_for('ssc_login_page'))

if __name__ == '__main__':
    app.run()

app = Flask(__name__)
app.config['TEMPLATE_AUTO_RELOAD'] = True